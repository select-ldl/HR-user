# ```python
import openpyxl
import mysql.connector

# 打开xlsx文件
wb = openpyxl.load_workbook('example.xlsx')

# 选择工作表
sheet = wb['Sheet1']

# 连接MySQL数据```python
import openpyxl
import mysql.connector

# 打开xlsx文件
wb = openpyxl.load_workbook('example.xlsx')

# 选择工作表
sheet = wb['Sheet1']

# 连接MySQL数据库
mydb = mysql.connector.connect(
  host="localhost",
  user="yourusername",
  password="yourpassword",
  database="yourdatabase"
)

# 创建游标对象
mycursor = mydb.cursor()

# 遍历行，生成SQL语句并插入到数据库
for row in sheet.iter_rows(min_row=2, values_only=True):
    sql = "INSERT INTO employees (name, age, salary) VALUES (%s, %s, %s)"
    val = (row[0], row[1], row[2])
    mycursor.execute(sql, val)

# 提交更改
mydb.commit()

# 输出插入的行数
print(mycursor.rowcount, "rows inserted.")
mydb = mysql.connector.connect(
  host="localhost",
  user="yourusername",
  password="yourpassword",
  database="yourdatabase"
)

# 创建游标对象
mycursor = mydb.cursor()

# 遍历行，生成SQL语句并插入到数据库
for row in sheet.iter_rows(min_row=2, values_only=True):
    sql = "INSERT INTO employees (name, age, salary) VALUES (%s, %s, %s)"
    val = (row[0], row[1], row[2])
    mycursor.execute(sql, val)

# 提交更改
mydb.commit()

# 输出插入的行数
print(mycursor.rowcount, "rows inserted.")